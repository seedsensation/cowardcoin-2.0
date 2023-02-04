from global_context import *
from openpyxl import Workbook,load_workbook
import os.path
from commands_folder.store_coins_file import *
from discord.utils import get
from time import sleep
from math import floor
async def vote(ctx,args):
    decrees = list(election.keys())
    if len(args) == 1:
        print(decrees)
        output = """Welcome to the Future of CowardCoin.
Welcome to Fair Play.
Starting now, each week, an Election shall be run.
The Law with the most Votes will become Law for the entire Server.
To Vote, type `!coin vote [LETTER] [NUMBER]` - e.g., `!coin vote A 12`, to put 12 Votes into """+str(election[decrees[0]][0])+""".
One Coin is One Vote.
Help Us decide the Future of CowardCoin.
The Future is Bright.
The Future is Fair Play.

***LAWS:***\n"""
        for item in decrees:
            output+=f"{item} - **{election[item][0]}**\n{election[item][1]}\n\n"

        await ctx.send(output)

    else:
        try:
            test = float(args[2])%1
        except ValueError:
            await ctx.send("NUMBER FAIL !!")
            test = 1

        if len(args[1]) == 1 and (test)==0:
            if int(args[2]) <= context[2][ctx.author.id][0]:
                if not os.path.exists("elections.xlsx"):
                    workbook = Workbook()
                    sheet = workbook.active
                    x = 0
                    decrees = list(election.keys())
                    for item in decrees:
                        x += 1
                        sheet[f"A{x}"] = item
                        sheet[f"B{x}"] = 0
                    workbook.save(filename="elections.xlsx")

                workbook = load_workbook(filename="elections.xlsx")
                sheet = workbook.active
                complete = False
                x = 0
                while True:
                    x += 1
                    if sheet[f"A{x}"].value == None:
                        break
                    else:
                        if str(sheet[f"A{x}"].value).lower() == (args[1]).lower():
                            sheet[f"B{x}"].value += int(args[2])
                            complete = True
                try:
                    workbook.save(filename="elections.xlsx")

                    if complete:
                        await ctx.send(f"Your vote for {args[1]} has been submitted successfully!")
                    else:
                        await ctx.send(f"{args[1]} is not a valid Law.")

                    context[2][ctx.author.id][0] -= int(args[2])
                    await savecoins(ctx)

                except PermissionError:
                    await ctx.send("An error occurred. Please try again in a moment.")



            else:
                await ctx.send("You can't afford that many votes!")



        else:
            await ctx.send("Sorry, that number of votes isn't valid!")

async def enact(ctx):
    check_role = get(ctx.guild.roles, name="admins")
    if check_role not in ctx.author.roles:
        await ctx.send("Admins only... sorry :(")
    else:
        result = ["A"]
        s = ""
        if len(result) > 1:
            s = "s"
        output = f"The Following Law{s} will be Enacted with Immediate Effect:\n\n"
        for item in result:
            output += f"**{item}** - ***{pastelection[item][0]}*** - 95.39% of the Vote\n{pastelection[item][1]}\n\n"

        await ctx.send(output)
        sleep(5)
        await eat_the_rich(ctx)


async def eat_the_rich(ctx):
    await ctx.send("Executing 'Eat the Rich'.")
    sleep(2.5)
    coins = context[2]
    coindict = {}
    for item in coins:
        if coins[item][0] != 0:
            coindict[item] = coins[item][0]
    sortedcoindict = sorted(coindict.items(), key=lambda x: x[1], reverse=True)
    recipientobject = bot.get_user(sortedcoindict[0][0])
    await ctx.send(f"{recipientobject.display_name} is the Richest Collector this week.\nThey have {sortedcoindict[0][1]} coins.")
    amount = floor(sortedcoindict[0][1]/len(coins))
    coins[recipientobject.id][0] -= (amount*len(coins))
    for user in coins:
        coins[user][0] += amount
    sleep(2.5)
    await ctx.send(f"Everyone receives {amount} coins from the Redistribution.")
    print(coins)
    context[2] = coins
    await savecoins(ctx)
